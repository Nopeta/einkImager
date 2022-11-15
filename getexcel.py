import os
import openpyxl
import datetime
# from openpyxl.utils import get_column_letter, column_index_from_string
path = os.path.dirname(__file__) + '/'
wb = openpyxl.load_workbook(path + 'data_test.xlsx')

today = datetime.datetime.today().strftime('%m/%d')
year = datetime.datetime.today().strftime('%Y')
# names = wb.sheetnames
# s1 = wb.active
s1 = wb[year]
# print(s1.title, s1.max_row, s1.max_column)


def get_values(sheet):
    arr = []                      # 第一層串列
    for row in sheet:
        arr2 = []                # 第二層串列
        if row[0].value == 'date':
            continue
        else:
            if row[0].value.strftime('%m/%d') == today:
                for column in row:
                    arr2.append(column.value)  # 寫入內容
                # arr.append(arr2)
                obj = {"date": arr2[0].strftime('%m/%d'), "start": arr2[1].strftime('%H:%M'),
                       "end": arr2[2].strftime('%H:%M'), "name": arr2[3], "use": arr2[4]}
                arr.append(obj)
            else:
                continue
    return arr


print(get_values(s1))       # 印出工作表 1 所有內容

# print(column_index_from_string('A'))    # 1
# print(column_index_from_string('AA'))   # 27
# print(get_column_letter(5))             # E
# print(get_column_letter(100))           # CV
